import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO


HEADER_FILL = PatternFill("solid", fgColor="0A1520")
HEADER_FONT = Font(bold=True, color="00D4FF", name="Calibri", size=11)
ALT_FILL    = PatternFill("solid", fgColor="060D16")


def style_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell            = ws.cell(row=1, column=col, value=h)
        cell.font       = HEADER_FONT
        cell.fill       = HEADER_FILL
        cell.alignment  = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22


def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)


def generate_excel_report(bookings, penalties, stats) -> bytes:
    wb = openpyxl.Workbook()

    # ── Sheet 1: All Bookings ──────────────────
    ws1 = wb.active
    ws1.title = "All Bookings"
    headers1  = ["Booking ID","User","Slot","Vehicle No","Type",
                 "Check-in","Duration (h)","Amount (₹)","Status","Penalty (₹)"]
    style_header(ws1, headers1)

    for i, b in enumerate(bookings, 2):
        row_data = [
            b.get("booking_id"), b.get("user"),
            f"S{b.get('slot_number',0):02d}",
            b.get("vehicle_number"), b.get("vehicle_type"),
            b.get("checkin_time"), b.get("duration"),
            b.get("amount"), b.get("status").upper(),
            b.get("penalty", 0)
        ]
        for col, val in enumerate(row_data, 1):
            cell       = ws1.cell(row=i, column=col, value=val)
            cell.font  = Font(color="E0F4FF", name="Calibri")
            if i % 2 == 0:
                cell.fill = ALT_FILL
    auto_width(ws1)

    # ── Sheet 2: Penalties ─────────────────────
    ws2      = wb.create_sheet("Penalties")
    headers2 = ["Booking ID","User","Slot","Original (₹)","Penalty (₹)","Pct %","Reason","Date"]
    style_header(ws2, headers2)

    for i, p in enumerate(penalties, 2):
        row_data = [
            p.get("booking_id"), p.get("user"),
            f"S{p.get('slot_number',0):02d}",
            p.get("original_amount"), p.get("penalty_amount"),
            f"{p.get('penalty_pct',0)}%",
            p.get("reason"), p.get("created_at")
        ]
        for col, val in enumerate(row_data, 1):
            cell      = ws2.cell(row=i, column=col, value=val)
            cell.font = Font(color="FF2D78", name="Calibri")
    auto_width(ws2)

    # ── Sheet 3: Summary ───────────────────────
    ws3      = wb.create_sheet("Summary")
    headers3 = ["Metric", "Value"]
    style_header(ws3, headers3)

    rows3 = [
        ("Total Bookings",    stats.get("total_bookings")),
        ("Total Revenue (₹)", stats.get("total_revenue")),
        ("Total Penalties (₹)", stats.get("total_penalties")),
        ("Avg Duration (hrs)", stats.get("avg_duration")),
        ("Cancellation Rate",  stats.get("cancellation_rate")),
    ]
    for i, (metric, val) in enumerate(rows3, 2):
        ws3.cell(row=i, column=1, value=metric).font = Font(color="00D4FF", bold=True)
        ws3.cell(row=i, column=2, value=val).font    = Font(color="00FF9D")
    auto_width(ws3)

    # ── Return bytes ───────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
